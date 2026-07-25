from datetime import datetime
import io
from fpdf import FPDF
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

# Configuración de página
st.set_page_config(
    page_title="Control de Entregas - SONAMEX", page_icon="📦", layout="wide"
)

# Estilo visual corporativo SONAMEX
PRIMARY_COLOR = "#00512D"

st.markdown(
    f"""
    <style>
    .main-header {{
        font-size: 24px;
        font-weight: bold;
        color: {PRIMARY_COLOR};
        text-align: center;
        margin-bottom: 20px;
    }}
    .stButton>button {{
        background-color: {PRIMARY_COLOR};
        color: white;
        border-radius: 5px;
        width: 100%;
    }}
    </style>
""",
    unsafe_allow_html=True,
)

# Base de datos simulada en memoria
if "db_entregas" not in st.session_state:
    st.session_state.db_entregas = pd.DataFrame(
        columns=[
            "id",
            "fecha_asignacion",
            "cliente",
            "direccion",
            "telefono",
            "estatus_entrega",
            "motivo",
            "comentarios",
            "fecha_actualizacion",
        ]
    )

# Control de Sesión de Usuario
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.usuario = ""
    st.session_state.rol = ""

# Pantalla de Login
if not st.session_state.autenticado:
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        try:
            st.image("logo_sonamex.png", width=220)
        except Exception:
            pass
        st.markdown(
            '<div class="main-header">🔐 Iniciar Sesión - SONAMEX</div>',
            unsafe_allow_html=True,
        )

        with st.form("login_form"):
            user_input = st.text_input("Usuario")
            pass_input = st.text_input("Contraseña", type="password")
            submit_login = st.form_submit_button("Entrar al Sistema")

            if submit_login:
                usuarios_validos = {
                    "oficina1": {
                        "pass": "sonamex2026",
                        "rol": "Oficina",
                        "nombre": "Personal de Oficina",
                    },
                    "repartor1": {
                        "pass": "ruta123",
                        "rol": "Campo",
                        "nombre": "Repartidor en Ruta",
                    },
                }

                if (
                    user_input in usuarios_validos
                    and usuarios_validos[user_input]["pass"] == pass_input
                ):
                    st.session_state.autenticado = True
                    st.session_state.usuario = usuarios_validos[user_input][
                        "nombre"
                    ]
                    st.session_state.rol = usuarios_validos[user_input]["rol"]
                    st.success("¡Bienvenido! Cargando sistema...")
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos.")
    st.stop()

# --- SESIÓN ACTIVA ---
df = st.session_state.db_entregas

try:
    st.sidebar.image("logo_sonamex.png", width=180)
except Exception:
    st.sidebar.markdown("### SONAMEX")

st.sidebar.markdown(f"👤 **Usuario:** {st.session_state.usuario}")
st.sidebar.markdown(f"🏷️ **Perfil:** {st.session_state.rol}")

if st.sidebar.button("Cerrar Sesión"):
    st.session_state.autenticado = False
    st.rerun()

st.markdown(
    '<div class="main-header">📦 Sistema de Control y Medición de Entregas - SONAMEX</div>',
    unsafe_allow_html=True,
)

if st.session_state.rol == "Oficina":
    menu = st.sidebar.selectbox(
        "Selecciona el Módulo",
        ["1. Oficina (Alta de Envíos)", "2. Campo (Repartidores)", "3. Reportes"],
    )
else:
    menu = "2. Campo (Repartidores)"
    st.sidebar.info(
        "Modo Repartidor activo: Solo visualizas el módulo de campo."
    )

# ==========================================
# MÓDULO 1: OFICINA
# ==========================================
if menu == "1. Oficina (Alta de Envíos)":
    st.subheader("🏢 Módulo de Atención a Clientes / Oficina")
    st.write(
        "Ingresa los datos del cliente y la dirección para enviar a ruta."
    )

    with st.form("form_alta"):
        col1, col2 = st.columns(2)
        with col1:
            cliente = st.text_input("Nombre del Cliente")
            telefono = st.text_input("Teléfono de Contacto")
        with col2:
            fecha_asig = st.date_input(
                "Fecha programada", value=datetime.today()
            )
            direccion = st.text_area("Dirección Completa")

        submit_alta = st.form_submit_button("Registrar Envío para Ruta")

        if submit_alta:
            if cliente and direccion:
                nuevo_id = len(df) + 1
                nuevo_registro = {
                    "id": nuevo_id,
                    "fecha_asignacion": str(fecha_asig),
                    "cliente": cliente,
                    "direccion": direccion,
                    "telefono": telefono,
                    "estatus_entrega": "Pendiente",
                    "motivo": "",
                    "comentarios": "",
                    "fecha_actualizacion": "",
                }
                st.session_state.db_entregas = pd.concat(
                    [df, pd.DataFrame([nuevo_registro])], ignore_index=True
                )
                st.success(
                    f"¡Envío para el cliente '{cliente}' registrado con éxito!"
                )
            else:
                st.error("Por favor completa al menos el nombre y la dirección.")

    st.markdown("---")
    st.subheader("📋 Envíos Actuales en Sistema")
    if not df.empty:
        st.dataframe(df, use_container_width=True)
    else:
        st.info("No hay envíos registrados todavía.")

# ==========================================
# MÓDULO 2: CAMPO (Repartidores)
# ==========================================
elif menu == "2. Campo (Repartidores)":
    st.subheader("📱 Módulo de Reparto en Campo")
    st.write("Selecciona la entrega pendiente para actualizar su estatus.")

    if df.empty or len(df[df["estatus_entrega"] == "Pendiente"]) == 0:
        st.info("No hay entregas pendientes por actualizar en este momento.")
    else:
        pendientes = df[df["estatus_entrega"] == "Pendiente"]
        cliente_seleccionado = st.selectbox(
            "Selecciona el Cliente a actualizar:",
            options=pendientes["cliente"].tolist(),
        )

        registro_idx = df[df["cliente"] == cliente_seleccionado].index[0]
        row = df.loc[registro_idx]

        st.markdown(f"**Dirección:** {row['direccion']}")
        st.markdown(f"**Teléfono:** {row['telefono']}")

        with st.form("form_campo"):
            estatus = st.radio(
                "¿Se entregó el producto?",
                ["Entregado con éxito", "No se entregó"],
            )

            motivo = ""
            if estatus == "No se entregó":
                motivo = st.selectbox(
                    "Motivo de no entrega",
                    [
                        "Cliente ausente",
                        "Negocio cerrado",
                        "Dirección incorrecta",
                        "Reprogramado por cliente",
                        "Otro",
                    ],
                )

            comentarios = st.text_area(
                "Comentarios adicionales (Ej. Pide entregar por la tarde)"
            )
            submit_campo = st.form_submit_button("Guardar Estatus de Entrega")

            if submit_campo:
                st.session_state.db_entregas.at[
                    registro_idx, "estatus_entrega"
                ] = estatus
                st.session_state.db_entregas.at[registro_idx, "motivo"] = motivo
                st.session_state.db_entregas.at[
                    registro_idx, "comentarios"
                ] = comentarios
                st.session_state.db_entregas.at[
                    registro_idx, "fecha_actualizacion"
                ] = str(datetime.now())
                st.success("¡Estatus de entrega actualizado correctamente!")
                st.rerun()

# ==========================================
# MÓDULO 3: REPORTES (Excel / PDF por Fechas)
# ==========================================
elif menu == "3. Reportes":
    st.subheader("📊 Módulo de Reportería y Exportación por Fechas")
    st.write(
        "Selecciona el rango de fechas para extraer y exportar el reporte."
    )

    if df.empty:
        st.warning("No hay datos disponibles para generar reportes.")
    else:
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            f_inicio = st.date_input("Fecha Inicial", value=datetime.today())
        with col_f2:
            f_fin = st.date_input("Fecha Final", value=datetime.today())

        # Filtrar datos por fecha seleccionada
        df["fecha_dt"] = pd.to_datetime(
            df["fecha_asignacion"]
        ).dt.date
        df_filtrado = df[
            (df["fecha_dt"] >= f_inicio) & (df["fecha_dt"] <= f_fin)
        ].drop(columns=["fecha_dt"])

        st.info(
            f"Se encontraron **{len(df_filtrado)}** registros para el rango seleccionado."
        )

        tipo_exportacion = st.radio(
            "Selecciona el formato de exportación:", ["Excel (.xlsx)", "PDF"]
        )

        if tipo_exportacion == "Excel (.xlsx)":
            if st.button("📥 Generar y Descargar Reporte en EXCEL"):
                output = io.BytesIO()
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Reporte de Entregas"

                try:
                    img = OpenpyxlImage("logo_sonamex.png")
                    img.width = 140
                    img.height = 45
                    ws.add_image(img, "B2")
                except Exception:
                    pass

                header_fill = PatternFill(
                    start_color="00512D", end_color="00512D", fill_type="solid"
                )
                header_font = Font(
                    name="Calibri", size=11, bold=True, color="FFFFFF"
                )
                border_thin = Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"),
                )

                ws.append([])
                ws.append([])
                ws.append([])
                ws.append(["REPORTE DE ENTREGAS - SONAMEX S.A. DE C.V."])
                ws.append(
                    [
                        f"Período: {f_inicio} al {f_fin} | Emitido: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    ]
                )
                ws.append([])

                headers = [
                    "ID",
                    "Fecha Asignación",
                    "Cliente",
                    "Dirección",
                    "Teléfono",
                    "Estatus",
                    "Motivo",
                    "Comentarios",
                    "Última Actualización",
                ]
                ws.append(headers)

                header_row_idx = ws.max_row
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=header_row_idx, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                for row_idx, row in df_filtrado.iterrows():
                    ws.append(list(row))
                    current_row = ws.max_row
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.border = border_thin
                        cell.alignment = Alignment(
                            horizontal="left", vertical="center"
                        )

                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(
                        max_len + 3, 12
                    )

                wb.save(output)
                processed_data = output.getvalue()

                st.download_button(
                    label="⬇️ Descargar Archivo Excel (.xlsx)",
                    data=processed_data,
                    file_name=f"Reporte_Entregas_{f_inicio}_al_{f_fin}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.success("¡Reporte en Excel generado con éxito!")

        else:
            if st.button("📥 Generar y Descargar Reporte en PDF"):
                pdf = FPDF(orientation="L", unit="mm", format="A4")
                pdf.add_page()

                # Cabecera PDF con colores Sonamex
                pdf.set_fill_color(0, 81, 45)  # Verde Sonamex
                pdf.rect(0, 0, 297, 20, "F")

                pdf.set_font("Arial", "B", 14)
                pdf.set_text_color(255, 255, 255)
                pdf.set_xy(10, 5)
                pdf.cell(
                    0,
                    10,
                    "REPORTE DE ENTREGAS - SONAMEX S.A. DE C.V.",
                    0,
                    1,
                    "L",
                )

                pdf.set_font("Arial", "", 10)
                pdf.set_text_color(50, 50, 50)
                pdf.set_xy(10, 25)
                pdf.cell(
                    0,
                    10,
                    f"Rango de Fechas: {f_inicio} al {f_fin} | Emitido: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    0,
                    1,
                    "L",
                )

                pdf.ln(5)

                # Tabla PDF
                pdf.set_font("Arial", "B", 8)
                pdf.set_fill_color(230, 235, 230)
                headers = [
                    "ID",
                    "Fecha",
                    "Cliente",
                    "Dirección",
                    "Teléfono",
                    "Estatus",
                    "Motivo",
                    "Comentarios",
                ]
                col_widths = [10, 22, 45, 75, 25, 30, 30, 40]

                for i, header in enumerate(headers):
                    pdf.cell(
                        col_widths[i], 8, header, 1, 0, "C", fill=True
                    )
                pdf.ln()

                pdf.set_font("Arial", "", 7)
                for _, row in df_filtrado.iterrows():
                    pdf.cell(col_widths[0], 7, str(row["id"]), 1, 0, "C")
                    pdf.cell(
                        col_widths[1], 7, str(row["fecha_asignacion"]), 1, 0, "C"
                    )
                    pdf.cell(
                        col_widths[2],
                        7,
                        str(row["cliente"])[:25],
                        1,
                        0,
                        "L",
                    )
                    pdf.cell(
                        col_widths[3],
                        7,
                        str(row["direccion"])[:45],
                        1,
                        0,
                        "L",
                    )
                    pdf.cell(
                        col_widths[4], 7, str(row["telefono"]), 1, 0, "C"
                    )
                    pdf.cell(
                        col_widths[5],
                        7,
                        str(row["estatus_entrega"]),
                        1,
                        0,
                        "C",
                    )
                    pdf.cell(col_widths[6], 7, str(row["motivo"])[:18], 1, 0, "L")
                    pdf.cell(
                        col_widths[7],
                        7,
                        str(row["comentarios"])[:25],
                        1,
                        0,
                        "L",
                    )
                    pdf.ln()

                # Corrección aplicada aquí para fpdf2 actual:
                pdf_output = bytes(pdf.output())

                st.download_button(
                    label="⬇️ Descargar Archivo PDF",
                    data=pdf_output,
                    file_name=f"Reporte_Entregas_{f_inicio}_al_{f_fin}.pdf",
                    mime="application/pdf",
                )
                st.success("¡Reporte en PDF generado con éxito!")
